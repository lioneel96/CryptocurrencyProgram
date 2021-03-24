# -*- coding: utf-8 -*-
import bs4
import requests
from openpyxl import load_workbook
import time
import os

try:
    from Tkinter import *   # for Python2
    from Tkinter import messagebox
except ImportError:
    from tkinter import *   # for Python3
    from tkinter import messagebox
    from tkinter import filedialog

#-----------VARIABLES GLOBALES------------
datos=[]
variables=[]
#entries=[]
monedas=["BTC","ETH","XRP","EOS","DOT1","MIOTA","LTC","ADA","LINK","XLM","XMR","OMG","NEO","EGLD"]
urls=[]
for i in range(len(monedas)):
    urls.append("https://finance.yahoo.com/quote/"+monedas[i]+"-USD")
    #print(urls[index])

#-----------FUNCIONES------------

def parsePrice():
    global urls
    r=[]
    r.clear()
    web_content=[]
    web_content.clear()
    price=[]
    price.clear()
    minmax_list=[]
    minmax_list.clear()
    day_min=[]
    day_min.clear()
    day_max=[]
    day_max.clear()
    for index in range(len(urls)):
        r.append( requests.get( urls[index] ) )
        web_content.append( bs4.BeautifulSoup(r[index].text,'lxml') )
        price.append(web_content[index].find_all('span',{'class':'Trsdu(0.3s) Fw(b) Fz(36px) Mb(-4px) D(ib)'})[0].text)
        minmax_list.append(web_content[index].find_all('td',{'data-reactid':'57'})[0].text)
        #print("Moneda: "+monedas[index]+" Precio:"+price[index]+" MIN-MAX:"+minmax_list[index])
        minmax=minmax_list[index].split(" - ")
        day_min.append(minmax[0])
        day_max.append(minmax[1])
    return price, day_min, day_max

def actualizaValores():
    price, day_min, day_max = parsePrice()
    output_list = convierteValores(price, day_min, day_max)
    calculaValores(output_list)
    status_text.set(time.strftime("%d/%m/%y - %H:%M:%S"))
    actualizaTabla()
#     modificaCeldas()
    
def convierteValores(price, day_min, day_max):
    output_list=[]
    output_list.clear()
    for index in range(len(price)): #-- ELIMINO LAS COMAS
        if("," in price[index]):
            value=price[index].split(",")
            value=value[0]+value[1]
        else:
            value=price[index]

        if("," in day_min[index]):
            value_min=day_min[index].split(",")
            value_min=value_min[0]+value_min[1]
        else:
            value_min=day_min[index]
        
        if("," in day_max[index]):
            value_max=day_max[index].split(",")
            value_max=value_max[0]+value_max[1]
        else:
            value_max=day_max[index]
            
        output_list.append([float(value),float(value_min),float(value_max)])
        #print(output_list[index])
    return output_list
                              
def actualizaExcel():
    global datos
    global path_text
    filename=path_text.get()
    wb=load_workbook(filename)
    ws=wb.worksheets[0]
    fila=0
    for index in range(len(datos)):
        for item in range(len(datos[index])-2):
            ws.cell(4+fila,item+3,float(datos[index][item]))
        fila+=1
        
    wb.save(filename)

def actualizaTabla():
    global datos
    for i in range(len(datos)):
        for j in range(5):
            variables[(5*i)+j].set(str(datos[i][j]))

def creaTabla():
    global variables
    global monedas
    for i in range(len(monedas)*5):
        variables.append( StringVar() )
        variables[i].set("0")
    for i in range(len(monedas)):
        for j in range(5):
            Label(miFrame,textvariable=variables[(5*i)+j],bd="5").grid(row=i+1,column=j+1,padx=5)


# def modificaCeldas():
#     for i in range(len(datos)):
#         if(datos[i][4]>3):
#             entries[(5*i)+4].config(background="green")
#         elif(datos[i][4]>2):
#             entries[(5*i)+4].config(background="limegreen")
#         elif(datos[i][4]>1):
#             entries[(5*i)+4].config(background="lightgreen")
#         elif(datos[i][4]>0):
#             entries[(5*i)+4].config(background="#c1fcc1")
#         elif(datos[i][4]>-1):
#             entries[(5*i)+4].config(background="lightcoral")
#         else:
#             entries[(5*i)+4].config(background="red")

def calculaValores(output_list):
    GP=[]
    P=[]
    GP.clear()
    P.clear()
    datos.clear()
    for i in range(len(output_list)):
        GP.append( round( (output_list[i][0]+output_list[i][1]+output_list[i][2])/3.0 , 3 ) )
        P.append( round( ((output_list[i][0]-GP[i])/GP[i])*100 , 3 ) )
        #print("GP:"+str(GP[i])+" P:"+str(P[i]))
        output_list[i].append(GP[i])
        output_list[i].append(P[i])
        datos.append(output_list[i])
        
def salir():
    rta=messagebox.askquestion("Salir", "¿Desea salir de la aplicación?")
    if rta=="yes" :
        raiz.destroy()

def cambiaPath():
    global path_text
    path_text.set(filedialog.askopenfilename(title="Abrir", initialdir=path_text.get(),filetypes=(("Ficheros de Excel","*xlsx"),("Todos los ficheros","*.*"))))


#-----------INTERFAZ GRÁFICA-------------
raiz=Tk()
raiz.title("Cryptocurrency values from Yahoo")
#raiz.iconbitmap("btc_icon.ico")

firstFrame=Frame(raiz)
#firstFrame.config(bg="red")
firstFrame.pack()
 
status_text=StringVar()
status_text.set("Never updated")
path_text=StringVar()
path_text.set(os.getcwd()+"\\trade.xlsx")
Label(firstFrame, text="Cryptocurrency values and data, by L&M").grid(row=0,column=0,padx=10,pady=5,columnspan=6)
Label(firstFrame, text="Last update: ").grid(row=1,column=0,pady=2,sticky="e")
Label(firstFrame, textvariable=status_text).grid(row=1,column=1,pady=2,sticky="w")
Label(firstFrame, text="Excel path: ").grid(row=2,column=0,pady=2,sticky="e")
Label(firstFrame, textvariable=path_text).grid(row=2,column=1,pady=2,sticky="w")
Button(firstFrame,text="Cambiar", width=10, command=cambiaPath).grid(row=2, column=2,pady=2)

miFrame=Frame(raiz)
#miFrame.config(bg="black")
miFrame.pack()

Label(miFrame, text="Coins", bd="10",bg="#FCD25A").grid(row=0,column=0,padx=10,pady=10)
Label(miFrame, text="ACTUAL",bd="10",bg="#FCD25A").grid(row=0,column=1,padx=10,pady=10)
Label(miFrame, text="MIN", bd="10",bg="#FCD25A").grid(row=0,column=2,padx=10,pady=10)
Label(miFrame, text="MAX",bd="10",bg="#FCD25A").grid(row=0,column=3,padx=10,pady=10)
Label(miFrame, text="Ghost Pivot",bd="10",bg="#FCD25A").grid(row=0,column=4,padx=10,pady=10)
Label(miFrame, text="% Pivot",bd="10",bg="#FCD25A").grid(row=0,column=5,padx=10,pady=10)

for i in range(len(monedas)):
    Label(miFrame, text=monedas[i], bd="5").grid(row=1+i,column=0)

creaTabla()

lastFrame=Frame(raiz)
#lastFrame.config(bg="blue")
lastFrame.pack()
Button(lastFrame,text="Actualizar valores",command=actualizaValores).grid(row=0, column=1,padx=10,pady=5)
Button(lastFrame,text="Actualizar Excel",command=actualizaExcel).grid(row=0, column=2,padx=10,pady=5)
Button(lastFrame,text="Salir", width=10,command=salir).grid(row=0, column=3,padx=10,pady=5)

raiz.mainloop()